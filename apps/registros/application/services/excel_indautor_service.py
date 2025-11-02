import os
import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Protection
from openpyxl.workbook.protection import WorkbookProtection
from django.db import connection


class ExcelIndautorService:
    def __init__(self):
        self.assets_path = os.path.join(
            os.path.dirname(__file__), "..", "..", "..", "..", "assets"
        )
        self.template_file = os.path.join(self.assets_path, "indautor.xlsx")
        self.output_file = self.template_file
        self.protection_password = "indautor2025"

        self.temas = {
            "Ramas": 3,
            "Medio de Ingreso": 4,
            "Estatus": 7,
            "Programa Educativo": 12,
            "Cuerpo Academico": 13,
            "Departamento": 11,
            "Sector": 2,
            "Sexo": 1,
            "TipInvestigador": 10,
        }

        self.column_order = [
            "Ramas",
            "Medio de Ingreso",
            "Ciudad de Origen",
            "CEPAT",
            "CIA",
            "Año Renovación",
            "Año CIA",
            "Estatus",
            "Programa Educativo",
            "Cuerpo Academico",
            "Departamento",
            "Sector",
            "Sexo",
            "TipInvestigador",
        ]

    def execute(self):
        if not os.path.exists(self.template_file):
            raise FileNotFoundError(f"No se encontró: {self.template_file}")

        wb = load_workbook(self.template_file, data_only=False, keep_links=True)
        self.fill_clasificaciones(wb)
        self.fill_sector_sheets(wb)
        self.protect_and_hide_clasificaciones(wb)

        if wb.security is None:
            wb.security = WorkbookProtection()

        wb.security.lockStructure = True
        wb.security.workbookPassword = self.protection_password
        wb.save(self.output_file)

        return {
            "mensaje": "Plantilla actualizada con todas las hojas ocultas y protegidas (incluida CLASIFICACIONES).",
            "archivo": self.output_file,
        }

    def fill_clasificaciones(self, wb):
        sheet_name = next(
            (name for name in wb.sheetnames if name.strip().lower() == "clasificaciones"),
            None,
        )
        if not sheet_name:
            raise KeyError("No se encontró la hoja 'CLASIFICACIONES'.")

        ws = wb[sheet_name]
        if hasattr(ws, "_tables") and ws._tables:
            for tname in list(ws._tables.keys()):
                del ws._tables[tname]

        data = {}
        with connection.cursor() as cursor:
            for col, id_tema in self.temas.items():
                if col == "Ramas":
                    cursor.execute(
                        """
                        SELECT nombre
                        FROM parametrizacion
                        WHERE id_tema = %s
                          AND id_param_padre = 45
                        ORDER BY id_param;
                        """,
                        [id_tema],
                    )
                else:
                    cursor.execute(
                        """
                        SELECT nombre
                        FROM parametrizacion
                        WHERE id_tema = %s
                        ORDER BY id_param;
                        """,
                        [id_tema],
                    )
                data[col] = [r[0] for r in cursor.fetchall()]

            cursor.execute("SELECT nombre FROM institucion ORDER BY id_institucion;")
            data["Ciudad de Origen"] = [r[0] for r in cursor.fetchall()]

            cursor.execute("SELECT nombre FROM cepat ORDER BY id_cepat;")
            data["CEPAT"] = [r[0] for r in cursor.fetchall()]

        current_year = datetime.date.today().year
        data["Año Renovación"] = ["No Aplica"] + [
            str(y) for y in range(current_year - 10, current_year + 52)
        ]

        max_len = max((len(v) for v in data.values()), default=0)
        for c_idx, col_name in enumerate(self.column_order, start=1):
            valores = data.get(col_name, [])
            valores += [""] * (max_len - len(valores))
            for r_idx, val in enumerate(valores, start=2):
                ws.cell(row=r_idx, column=c_idx).value = val

    def fill_sector_sheets(self, wb):
        with connection.cursor() as cursor:
            cursor.execute(
                """
                SELECT id_param, nombre
                FROM parametrizacion
                WHERE id_tema = 2
                ORDER BY id_param;
                """
            )
            sectores = cursor.fetchall()

            for id_sector, nombre_sector in sectores:
                sheet_name = nombre_sector.upper().strip()
                if sheet_name in wb.sheetnames:
                    ws_sector = wb[sheet_name]
                    if hasattr(ws_sector, "_tables") and ws_sector._tables:
                        for tname in list(ws_sector._tables.keys()):
                            del ws_sector._tables[tname]
                    for row in ws_sector.iter_rows():
                        for cell in row:
                            cell.value = None
                else:
                    ws_sector = wb.create_sheet(sheet_name)

                cursor.execute(
                    """
                    SELECT id_param, nombre
                    FROM parametrizacion
                    WHERE id_param_padre = %s
                    ORDER BY id_param;
                    """,
                    [id_sector],
                )
                subsectores = cursor.fetchall()
                headers = [s[1] for s in subsectores]

                for c_idx, header in enumerate(headers, start=1):
                    ws_sector.cell(row=1, column=c_idx).value = header

                for c_idx, (id_subsector, _) in enumerate(subsectores, start=1):
                    cursor.execute(
                        """
                        SELECT nombre
                        FROM parametrizacion
                        WHERE id_param_padre = %s
                        ORDER BY id_param;
                        """,
                        [id_subsector],
                    )
                    actividades = [r[0] for r in cursor.fetchall()]
                    for r_idx, val in enumerate(actividades, start=2):
                        ws_sector.cell(row=r_idx, column=c_idx).value = val

                for i, header in enumerate(headers, start=1):
                    col_letter = get_column_letter(i)
                    ws_sector.column_dimensions[col_letter].width = max(len(header) + 5, 25)

                ws_sector.sheet_state = "veryHidden"
                ws_sector.protection.set_password(self.protection_password)
                ws_sector.protection.enable()

    def protect_and_hide_clasificaciones(self, wb):
        sheet_name = next(
            (n for n in wb.sheetnames if n.strip().lower() == "clasificaciones"),
            None,
        )
        if not sheet_name:
            return

        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                cell.protection = Protection(locked=True, hidden=False)

        ws.protection.set_password(self.protection_password)
        ws.protection.enable()
        ws.protection.format_cells = False
        ws.protection.insert_columns = False
        ws.protection.insert_rows = False
        ws.protection.delete_columns = False
        ws.protection.delete_rows = False
        ws.protection.sort = False
        ws.protection.autoFilter = False
        ws.protection.objects = True
        ws.protection.scenarios = True
        ws.sheet_state = "veryHidden"
