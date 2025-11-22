import pandas as pd
from io import BytesIO
from django.http import FileResponse
from rest_framework.response import Response
from rest_framework.decorators import action
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.marker import DataPoint
from openpyxl.chart.label import DataLabelList

def dataframe_to_styled_excel_bytes(
        df: pd.DataFrame,
        sheet_name: str = "Sheet1",
        table_name: str = "Table1",
        column_mapping: dict = None,
        report_title: str = None,
        add_pie_chart: bool = False,
        chart_labels_col_idx: int = 1,
        chart_data_col_idx: int = 2
) -> BytesIO:
    """
    Genera un Excel estilizado con opción a gráfico de anillo (Doughnut)
    con tamaño ajustado (aprox 500x500px).
    """
    output = BytesIO()

    if column_mapping:
        df = df.rename(columns=column_mapping)

    header_rows_count = 4
    start_row_idx = header_rows_count

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name, startrow=start_row_idx)
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]
        max_col = df.shape[1] if df.shape[1] > 0 else 1
        last_col_letter = get_column_letter(max_col)
        header_colors = ["5B9BD5", "9DC3E6", "BDD7EE", "BDD7EE"]
        headers_text = [
            "TECNOLÓGICO NACIONAL DE MÉXICO",
            "SECRETARÍA DE EXTENSIÓN Y VINCULACIÓN",
            "DIRECCIÓN DE VINCULACIÓN E INTERCAMBIO ACADEMICO"
        ]
        for i, text in enumerate(headers_text):
            row_num = i + 1
            cell = worksheet[f"A{row_num}"]
            cell.value = text
            cell.font = Font(bold=True, size=12, color="000000")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill(start_color=header_colors[i], end_color=header_colors[i], fill_type="solid")
            worksheet.merge_cells(f"A{row_num}:{last_col_letter}{row_num}")

        row_num_title = 4
        title_text = report_title.upper() if report_title else "REPORTE GENERAL"
        cell_title = worksheet[f"A{row_num_title}"]
        cell_title.value = title_text
        cell_title.font = Font(bold=True, size=14, color="000000")
        cell_title.alignment = Alignment(horizontal="center", vertical="center")
        cell_title.fill = PatternFill(start_color=header_colors[3], end_color=header_colors[3], fill_type="solid")
        worksheet.merge_cells(f"A{row_num_title}:{last_col_letter}{row_num_title}")

        table_start_row = start_row_idx + 1
        table_end_row = table_start_row + df.shape[0]
        table_ref = f"A{table_start_row}:{last_col_letter}{table_end_row}"

        tab = Table(displayName=table_name, ref=table_ref)
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False,
                               showRowStripes=True, showColumnStripes=False)
        tab.tableStyleInfo = style
        try:
            worksheet.add_table(tab)
        except ValueError:
            pass

        for col_idx in range(1, max_col + 1):
            col_letter = get_column_letter(col_idx)
            worksheet.column_dimensions[col_letter].width = 25
            worksheet.cell(row=table_start_row, column=col_idx).alignment = Alignment(horizontal="center",
                                                                                      vertical="center")

        for idx, col in enumerate(df.columns, start=1):
            if pd.api.types.is_integer_dtype(df[col].dtype):
                fmt = '#,##0'
            elif pd.api.types.is_float_dtype(df[col].dtype):
                fmt = '#,##0.00'
            else:
                fmt = None
            if fmt:
                for row in range(table_start_row + 1, table_end_row + 1):
                    c = worksheet.cell(row=row, column=idx)
                    c.number_format = fmt
                    c.alignment = Alignment(horizontal='center')

        worksheet.freeze_panes = f"A{table_start_row + 1}"

        if add_pie_chart and df.shape[0] > 0:
            chart = PieChart()
            chart.height = 16.0
            chart.width = 18.0

            chart.doughnut = True
            chart.holeSize = 65
            try:
                col_data_index = chart_data_col_idx - 1
                total_value = df.iloc[:, col_data_index].sum()
                chart.title = f"Total de registros {int(total_value):}"
            except:
                chart.title = " "

            data_ref = Reference(worksheet, min_col=chart_data_col_idx, min_row=table_start_row, max_row=table_end_row)
            labels_ref = Reference(worksheet, min_col=chart_labels_col_idx, min_row=table_start_row + 1,
                                   max_row=table_end_row)

            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(labels_ref)

            chart.dataLabels = DataLabelList()
            chart.dataLabels.showPercent = True
            chart.dataLabels.showVal = False
            chart.dataLabels.showCatName = False
            chart.legend.position = 'b'

            nice_colors = [
                "A3C4F3",  # Azul pastel suave
                "F6C6EA",  # Rosa pastel
                "A9E5BB",  # Verde menta suave
                "F9A875",  # Naranja pastel
                "F7E28B",  # Amarillo pastel suave
                "C5B8F1",  # Lavanda pastel
                "FFD8A8",  # Durazno claro
                "B2DFFB",  # Celeste pastel
                "F4A7B9",  # Rosa frambuesa
                "D7F2BA",  # Verde lima pastel
            ]
            series = chart.series[0]
            for i in range(len(df)):
                pt = DataPoint(idx=i)
                pt.graphicalProperties.solidFill = nice_colors[i % len(nice_colors)]
                series.dPt.append(pt)

            chart_start_col_letter = get_column_letter(max_col + 2)
            position = f"{chart_start_col_letter}{header_rows_count + 2}"
            worksheet.add_chart(chart, position)

    output.seek(0)
    return output


def dataframe_to_excel_registros_mes(
        df: pd.DataFrame,
        sheet_name: str = "Sheet1",
        table_name: str = "Table1",
        report_title: str = None,
        chart_labels_col_idx: int = 1,
        chart_data_col_idx: int = 2
) -> BytesIO:
    """
    Función ESPECÍFICA para el reporte de registros por mes.
    - Elimina la leyenda (que ocupaba mucho espacio).
    - Pone el nombre de la categoría y el porcentaje directo en el gráfico.
    - Aumenta el tamaño del gráfico.
    """
    output = BytesIO()

    header_rows_count = 4
    start_row_idx = header_rows_count

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name, startrow=start_row_idx)
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]
        max_col = df.shape[1] if df.shape[1] > 0 else 1
        last_col_letter = get_column_letter(max_col)

        # --- ESTILOS DE CABECERA (Igual que tu función original) ---
        header_colors = ["5B9BD5", "9DC3E6", "BDD7EE", "BDD7EE"]
        headers_text = [
            "TECNOLÓGICO NACIONAL DE MÉXICO",
            "SECRETARÍA DE EXTENSIÓN Y VINCULACIÓN",
            "DIRECCIÓN DE VINCULACIÓN E INTERCAMBIO ACADEMICO"
        ]
        for i, text in enumerate(headers_text):
            row_num = i + 1
            cell = worksheet[f"A{row_num}"]
            cell.value = text
            cell.font = Font(bold=True, size=12, color="000000")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill(start_color=header_colors[i], end_color=header_colors[i], fill_type="solid")
            worksheet.merge_cells(f"A{row_num}:{last_col_letter}{row_num}")

        row_num_title = 4
        title_text = report_title.upper() if report_title else "REPORTE"
        cell_title = worksheet[f"A{row_num_title}"]
        cell_title.value = title_text
        cell_title.font = Font(bold=True, size=14, color="000000")
        cell_title.alignment = Alignment(horizontal="center", vertical="center")
        cell_title.fill = PatternFill(start_color=header_colors[3], end_color=header_colors[3], fill_type="solid")
        worksheet.merge_cells(f"A{row_num_title}:{last_col_letter}{row_num_title}")

        # --- TABLA ---
        table_start_row = start_row_idx + 1
        table_end_row = table_start_row + df.shape[0]
        table_ref = f"A{table_start_row}:{last_col_letter}{table_end_row}"

        tab = Table(displayName=table_name, ref=table_ref)
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False,
                               showRowStripes=True)
        tab.tableStyleInfo = style
        try:
            worksheet.add_table(tab)
        except ValueError:
            pass

        # Ajuste de columnas
        for col_idx in range(1, max_col + 1):
            col_letter = get_column_letter(col_idx)
            worksheet.column_dimensions[col_letter].width = 30  # Un poco más ancho para que se lea "Enero - IMPI"

        # --- GRÁFICO PERSONALIZADO ---
        if df.shape[0] > 0:
            chart = PieChart()
            chart.height = 20.0  # Más alto
            chart.width = 25.0  # Más ancho para que quepan las etiquetas
            chart.doughnut = True
            chart.holeSize = 60  # Agujero un poco más pequeño para dar más espacio al color

            try:
                col_data_index = chart_data_col_idx - 1
                total_value = df.iloc[:, col_data_index].sum()
                chart.title = f"Total de registros: {int(total_value)}"
            except:
                chart.title = ""

            data_ref = Reference(worksheet, min_col=chart_data_col_idx, min_row=table_start_row, max_row=table_end_row)
            labels_ref = Reference(worksheet, min_col=chart_labels_col_idx, min_row=table_start_row + 1,
                                   max_row=table_end_row)

            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(labels_ref)

            # --- CONFIGURACIÓN CLAVE PARA QUE SE VEA BIEN ---
            chart.dataLabels = DataLabelList()
            chart.dataLabels.showPercent = True  # Muestra %
            chart.dataLabels.showCatName = True  # Muestra "Enero - IMPI"
            chart.dataLabels.showVal = False  # Oculta el número absoluto para ahorrar espacio

            # Opcional: Separador (depende de la versión de Excel del usuario, pero ayuda)
            chart.dataLabels.separator = "\n"

            chart.legend = None  # ADIÓS LEYENDA (Esto limpia el desastre)

            # Colores
            nice_colors = [
                "A3C4F3", "F6C6EA", "A9E5BB", "F9A875", "F7E28B", "C5B8F1",
                "FFD8A8", "B2DFFB", "F4A7B9", "D7F2BA"
            ]
            series = chart.series[0]
            for i in range(len(df)):
                pt = DataPoint(idx=i)
                pt.graphicalProperties.solidFill = nice_colors[i % len(nice_colors)]
                series.dPt.append(pt)

            chart_start_col_letter = get_column_letter(max_col + 2)
            position = f"{chart_start_col_letter}{header_rows_count + 2}"
            worksheet.add_chart(chart, position)

    output.seek(0)
    return output